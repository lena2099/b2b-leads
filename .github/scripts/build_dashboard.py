#!/usr/bin/env python3
"""Build dashboard: Excel, bilingual HTML, Obsidian vault from JSON data."""
import json, os
from datetime import datetime, timezone
from pathlib import Path

OUTPUT_DIR = Path("output")
IND_META = {
    "robotics": {"emoji":"🤖","cn":"机器人&自动化","en":"Robotics & Automation"},
    "energy_storage": {"emoji":"🔋","cn":"储能&电池","en":"Energy Storage & Battery"},
    "solar_pv": {"emoji":"☀️","cn":"光伏组件&逆变器","en":"Solar PV & Inverters"},
    "medical_device": {"emoji":"🏥","cn":"医疗器械&设备","en":"Medical Devices"},
    "laser_equipment": {"emoji":"🔩","cn":"激光设备&精密加工","en":"Laser Equipment"},
    "ev_charger": {"emoji":"🚗","cn":"新能源汽车&充电桩","en":"EV & Charging"},
    "construction_machinery": {"emoji":"🏗️","cn":"工程机械&农业机械","en":"Construction & Ag Machinery"},
    "telecom_iot": {"emoji":"📡","cn":"通信设备&物联网","en":"Telecom & IoT"},
}

all_data = {}
for ik in IND_META:
    fp = OUTPUT_DIR / f"{ik}_leads.json"
    if fp.exists():
        all_data[ik] = json.loads(fp.read_text())
    else:
        all_data[ik] = {"leads": []}

now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
all_leads = [l for v in all_data.values() for l in v["leads"]]
total = len(all_leads)
high = sum(1 for l in all_leads if l.get("relevance", 0) >= 70)

# ── EXCEL ──
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️ openpyxl not installed, skipping Excel")

def make_excel(leads, title, filename):
    wb = Workbook(); ws = wb.active; ws.title = title[:30]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfl = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    gf = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yf = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    rf = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cy = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")

    headers = ["Company","Website","Founded","Headquarters","Employees","Revenue",
               "Business","Procurement Needs","Partner Profile","Key Markets",
               "China Presence","Competitors","Market Share","Recent News",
               "Buyer Type","Country","Size","Relevance","Why"]
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h); c.font = hf; c.fill = hfl; c.alignment = ha; c.border = tb

    for i, l in enumerate(sorted(leads, key=lambda x: x.get("relevance",0), reverse=True), 2):
        vals = [
            l.get("company_name",""), l.get("website",""), l.get("founded",""), l.get("headquarters",""),
            l.get("employees",""), l.get("annual_revenue",""), l.get("business_description",""),
            l.get("procurement_needs",""), l.get("partner_profile",""), l.get("key_markets",""),
            l.get("china_presence",""), l.get("competitors",""), l.get("market_share",""),
            l.get("recent_news",""), l.get("buyer_type",""), l.get("country",""), l.get("size",""),
            l.get("relevance",0), l.get("why","")
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val); c.border = tb; c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 18:
                if val >= 70: c.fill = gf
                elif val >= 40: c.fill = yf
                else: c.fill = rf
            if col == 11 and isinstance(val, str) and not val.startswith("暂无") and val:
                c.fill = cy

    for col, w in enumerate([28,25,6,20,10,10,35,35,35,25,22,25,18,30,20,12,12,8,25], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = ws.dimensions; ws.freeze_panes = "A2"
    wb.save(str(OUTPUT_DIR / filename))

for ik, meta in IND_META.items():
    leads = all_data[ik]["leads"]
    if leads:
        make_excel(leads, meta["en"], f"{ik}_buyers.xlsx")
print(f"✅ {len(IND_META)} Excel files")

# ── BILINGUAL HTML ──
def make_html(lang):
    t = {
        "zh": {"title":"🌍 B2B海外买家情报","sub":"8行业·7×24自动挖掘","total":"总买家","high":"高优先级","ind":"行业","upd":"更新","co":"公司","ty":"类型","sc":"评分","sz":"规模","ct":"国家","why":"采购动机","cp":"中国业务","comp":"竞争对手","mkt":"市场份额","news":"近期动态","rub":"📊 评分标准","def":"定义","r1":"顶级·已知采购中国产品","r2":"强匹配·成本敏感","r3":"中等·潜在买家","r4":"弱匹配·偶发","ft":"B2B Lead Miner · DeepSeek AI · 每小时更新"},
        "en": {"title":"🌍 B2B Lead Miner","sub":"8 Industries · 7×24 Mining","total":"Total","high":"High Priority","ind":"Industries","upd":"Updated","co":"Company","ty":"Type","sc":"Score","sz":"Size","ct":"Country","why":"Why Buy","cp":"China Presence","comp":"Competitors","mkt":"Market Share","news":"Recent News","rub":"📊 Rubric","def":"Definition","r1":"Top-tier·Known China Buyer","r2":"Strong Fit·Price Sensitive","r3":"Moderate·Possible","r4":"Weak·Occasional","ft":"B2B Lead Miner · DeepSeek AI · Updated hourly"},
    }
    tx = t[lang]
    html = f'''<!DOCTYPE html><html lang="{lang}"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>{tx['title']}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#333}}
.hdr{{background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;padding:30px 40px;display:flex;justify-content:space-between;align-items:center}}
.hdr h1{{font-size:1.8em}}.hdr p{{color:#8892b0}}
.ls a{{padding:8px 18px;border-radius:20px;border:1px solid #8892b0;color:#8892b0;text-decoration:none;font-size:.85em;margin-left:8px}}
.ls a.ac{{background:#7b61ff;color:#fff;border-color:#7b61ff}}
.st{{display:flex;gap:20px;padding:20px 40px;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
.sti{{text-align:center;flex:1}}.stn{{font-size:2em;font-weight:700;color:#2B5797}}.stl{{font-size:.8em;color:#888;text-transform:uppercase}}
.ct{{max-width:1700px;margin:20px auto;padding:0 40px}}
.sec{{margin-bottom:30px}}.sec h2{{font-size:1.1em;color:#2B5797;margin:12px 0;padding-bottom:8px;border-bottom:2px solid #2B5797}}
.nv{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px;padding:14px;background:#fff;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
.nv a{{padding:5px 12px;border-radius:14px;background:#f0f2f5;color:#2B5797;text-decoration:none;font-size:.8em}}
.nv a:hover{{background:#2B5797;color:#fff}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.06);font-size:.78em}}
th{{background:#2B5797;color:#fff;padding:8px 10px;text-align:left;font-size:.75em}}
td{{padding:6px 10px;border-bottom:1px solid #e8e8e8}}tr:hover{{background:#f7f9fc}}
.hi{{background:#d4edda;color:#155724;padding:2px 6px;border-radius:3px;font-weight:600;font-size:.82em}}
.md{{background:#fff3cd;color:#856404;padding:2px 6px;border-radius:3px;font-weight:600;font-size:.82em}}
.lo{{background:#f8d7da;color:#721c24;padding:2px 6px;border-radius:3px;font-weight:600;font-size:.82em}}
.yb{{background:#d5e8d4;color:#2d6a2d;padding:2px 6px;border-radius:3px;font-size:.78em;font-weight:600}}
footer{{text-align:center;color:#999;padding:30px;font-size:.8em}}
</style></head><body>
<div class="hdr"><div><h1>{tx['title']}</h1><p>{tx['sub']}</p></div>
<div class="ls"><a href="dashboard.html" class="{'ac' if lang=='zh' else ''}">中文</a><a href="dashboard_en.html" class="{'ac' if lang=='en' else ''}">English</a></div></div>
<div class="st"><div class="sti"><div class="stn">{total}</div><div class="stl">{tx['total']}</div></div><div class="sti"><div class="stn">{high}</div><div class="stl">{tx['high']}</div></div><div class="sti"><div class="stn">8</div><div class="stl">{tx['ind']}</div></div><div class="sti"><div class="stn">{now}</div><div class="stl">{tx['upd']}</div></div></div>
<div class="ct"><div class="nv">'''
    for ik, meta in IND_META.items():
        html += f'<a href="#{ik}">{meta["emoji"]} {meta["cn"] if lang=="zh" else meta["en"]}</a>'
    html += f'''</div>
<div class="sec"><h3>{tx['rub']}</h3><table><tr><th>{tx['sc']}</th><th>{tx['def']}</th></tr><tr><td>90-95</td><td>{tx['r1']}</td></tr><tr><td>75-89</td><td>{tx['r2']}</td></tr><tr><td>55-74</td><td>{tx['r3']}</td></tr><tr><td>40-54</td><td>{tx['r4']}</td></tr></table></div>'''
    
    for ik, meta in IND_META.items():
        leads = all_data[ik]["leads"]
        label = f'{meta["emoji"]} {meta["cn"] if lang=="zh" else meta["en"]}'
        s = sorted(leads, key=lambda x: x.get("relevance",0), reverse=True)
        html += f'<div class="sec" id="{ik}"><h2>{label} ({len(s)})</h2><table><tr><th>{tx["co"]}</th><th>{tx["ct"]}</th><th>{tx["ty"]}</th><th>{tx["sc"]}</th><th>{tx["sz"]}</th><th>{tx["comp"]}</th><th>{tx["mkt"]}</th><th>{tx["cp"]}</th><th>{tx["news"]}</th></tr>'
        for l in s:
            sc = l.get("relevance",0)
            cl = "hi" if sc>=70 else ("md" if sc>=40 else "lo")
            cp = l.get("china_presence","")
            cph = f'<span class="yb">{cp[:60]}</span>' if cp and cp not in ("暂无","无","") else '<span style="color:#999;font-size:.78em">—</span>'
            news = l.get("recent_news","")[:80]
            html += f'<tr><td><strong>{l.get("company_name","")}</strong></td><td>{l.get("country","")}</td><td>{l.get("buyer_type","")}</td><td><span class="{cl}">{sc}</span></td><td>{l.get("size","")}</td><td style="font-size:.75em">{l.get("competitors","")[:50]}</td><td style="font-size:.75em">{l.get("market_share","")[:50]}</td><td>{cph}</td><td style="font-size:.75em">{news}</td></tr>'
        html += "</table></div>"
    html += f'</div><footer>{tx["ft"]}</footer></body></html>'
    return html

with open(OUTPUT_DIR / "dashboard.html", "w") as f: f.write(make_html("zh"))
with open(OUTPUT_DIR / "dashboard_en.html", "w") as f: f.write(make_html("en"))
print("✅ Bilingual dashboard done")
print(f"\n✨ Build complete. {total} buyers across 8 industries.")
