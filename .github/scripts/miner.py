#!/usr/bin/env python3
"""
B2B Lead Miner — Automated overseas buyer discovery for Chinese exporters.
Targets: Robotics + Energy Storage
Sources: Trade shows, industry directories, Google search, LinkedIn matching
Outputs: Excel (.xlsx) + HTML dashboard
Runs: Every 6 hours via GitHub Actions
"""
import json, os, re, sys, hashlib, time
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote

API_KEY = os.environ["DEEPSEEK_API_KEY"]
DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")
STATE_FILE = DATA_DIR / "state.json"
SEEN_FILE = DATA_DIR / "seen_domains.json"

# ── INDUSTRY CONFIG ──────────────────────────────────────
INDUSTRIES = {
    "robotics": {
        "name": "Robotics & Automation",
        "keywords": [
            "industrial robot", "collaborative robot", "warehouse automation",
            "AGV manufacturer", "robotics integrator", "factory automation",
            "robot distributor", "automation solution provider", "cobot reseller",
            "logistics automation", "pick and place robot", "welding robot",
            "palletizing robot", "robotic arm buyer", "industrial automation company"
        ],
        "target_regions": ["United States", "Germany", "Japan", "South Korea", "Canada", "UK", "France", "Italy", "Netherlands", "Australia"],
        "buyer_profiles": [
            "Automotive manufacturers & Tier 1 suppliers",
            "Electronics manufacturers (PCB assembly, semiconductor)",
            "Warehouse & logistics operators (3PL)",
            "Food & beverage processing plants",
            "Metal fabrication & welding shops",
            "Pharmaceutical manufacturers",
            "Plastics & injection molding companies",
        ]
    },
    "energy_storage": {
        "name": "Energy Storage & Battery",
        "keywords": [
            "energy storage system", "battery storage", "solar installer",
            "utility scale battery", "BESS provider", "energy storage integrator",
            "solar + storage installer", "EV charging network", "microgrid developer",
            "renewable energy developer", "power utility", "commercial solar provider",
            "residential battery installer", "grid operator", "energy management company"
        ],
        "target_regions": ["United States", "Germany", "UK", "Australia", "Japan", "Netherlands", "Spain", "Italy", "Canada", "Sweden"],
        "buyer_profiles": [
            "Utility companies (power generation & distribution)",
            "Solar installation companies (residential & commercial)",
            "Renewable energy project developers",
            "EV charging network operators",
            "Commercial real estate developers",
            "Industrial facility managers",
            "Microgrid developers",
            "Energy retailers & aggregators",
        ]
    }
}

# ── DATA SOURCES ─────────────────────────────────────────
TRADE_SHOWS = {
    "ces": {
        "name": "CES (Consumer Electronics Show)",
        "url": "https://www.ces.tech/exhibitor-directory.aspx",
        "industries": ["robotics"],
        "country": "United States",
    },
    "hannover_messe": {
        "name": "Hannover Messe",
        "url": "https://www.hannovermesse.de/en/expo/exhibitor-short-index/",
        "industries": ["robotics"],
        "country": "Germany",
    },
    "mwc": {
        "name": "MWC Barcelona",
        "url": "https://www.mwcbarcelona.com/exhibitors",
        "industries": ["robotics"],
        "country": "Spain",
    },
    "intersolar": {
        "name": "Intersolar Europe",
        "url": "https://www.intersolar.de/exhibitors-products/",
        "industries": ["energy_storage"],
        "country": "Germany",
    },
    "spi": {
        "name": "Solar Power International",
        "url": "https://www.solarpowerinternational.com/exhibitor-directory/",
        "industries": ["energy_storage"],
        "country": "United States",
    },
    "automatica": {
        "name": "Automatica (Robotics Fair)",
        "url": "https://automatica-munich.com/en/exhibitors/",
        "industries": ["robotics"],
        "country": "Germany",
    },
    "ees_europe": {
        "name": "ees Europe (Energy Storage)",
        "url": "https://www.ees-europe.com/exhibitors/",
        "industries": ["energy_storage"],
        "country": "Germany",
    },
}

GOOGLE_SEARCH_QUERIES = {
    "robotics": [
        'site:linkedin.com/company "industrial automation" "United States"',
        'site:linkedin.com/company "robotics integrator" "Germany"',
        '"warehouse automation solutions" "contact" site:linkedin.com',
        '"factory automation" distributor North America',
        '"collaborative robot" reseller OR distributor',
        'top robotics companies in {country} site:linkedin.com/company',
    ],
    "energy_storage": [
        'site:linkedin.com/company "energy storage" "solar" "United States"',
        'site:linkedin.com/company "battery storage" installer',
        '"utility scale battery" project developer contact',
        '"solar + storage" installer California OR Texas OR Florida',
        '"BESS" energy storage company Germany OR UK',
        'renewable energy project developer {country} site:linkedin.com',
    ],
}

# ── STATE MANAGEMENT ─────────────────────────────────────
def load_state():
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {"runs": 0, "total_leads": 0, "last_run": None}

def save_state(state):
    state["runs"] += 1
    state["last_run"] = datetime.now(timezone.utc).isoformat()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2))

def load_seen():
    if SEEN_FILE.exists():
        return set(json.loads(SEEN_FILE.read_text()))
    return set()

def save_seen(seen):
    SEEN_FILE.write_text(json.dumps(sorted(seen)))

def is_new(domain: str, seen: set) -> bool:
    return domain.lower() not in seen


# ── LLM ──────────────────────────────────────────────────
def call_deepseek(messages, max_tokens=1500, temperature=0.3):
    body = json.dumps({
        "model": "deepseek-chat",
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }).encode()
    req = Request("https://api.deepseek.com/chat/completions",
                  data=body,
                  headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"})
    resp = json.loads(urlopen(req, timeout=120).read())
    return resp["choices"][0]["message"]["content"].strip()


# ── SCRAPER: Trade Show Exhibitor Lists ─────────────────
def scrape_trade_show(show: dict, industry: str) -> list[dict]:
    """Scrape exhibitor list from a trade show website. Returns list of leads."""
    print(f"  🎪 {show['name']}...")
    leads = []
    try:
        url = show["url"]
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; B2BLeads/1.0)"})
        html = urlopen(req, timeout=20).read().decode("utf-8", errors="replace")

        # Extract company names from exhibitor pages (generic patterns)
        # Most trade show sites use similar patterns
        patterns = [
            r'<a[^>]*class="[^"]*exhibitor[^"]*"[^>]*>([^<]+)</a>',
            r'<h3[^>]*class="[^"]*company[^"]*"[^>]*>([^<]+)</h3>',
            r'data-company-name="([^"]+)"',
            r'<span[^>]*class="[^"]*exhibitor-name[^"]*"[^>]*>([^<]+)</span>',
            r'<div[^>]*class="[^"]*exhibitor[^"]*"[^>]*>\s*<h[234][^>]*>([^<]+)</h[234]>',
        ]
        companies = set()
        for pattern in patterns:
            matches = re.findall(pattern, html, re.IGNORECASE)
            for m in matches:
                name = re.sub(r'<[^>]+>', '', m).strip()
                if len(name) > 2 and len(name) < 150 and not name.startswith("<"):
                    companies.add(name)

        for company in companies:
            leads.append({
                "company_name": company,
                "source": show["name"],
                "source_type": "trade_show",
                "industry": industry,
                "country": show.get("country", "Unknown"),
                "found_at": datetime.now(timezone.utc).isoformat(),
            })

        print(f"     Found {len(leads)} exhibitors")
    except Exception as e:
        print(f"     ⚠️ Failed: {e}")

    return leads


# ── SCRAPER: Google Search → LinkedIn Companies ──────────
def scrape_google_linkedin(industry: str) -> list[dict]:
    """Search Google for LinkedIn company pages in target industries."""
    print(f"  🔍 Google → LinkedIn search for {industry}...")
    leads = []
    config = INDUSTRIES[industry]
    queries = GOOGLE_SEARCH_QUERIES.get(industry, [])

    for region in config["target_regions"][:5]:  # Limit to avoid rate limits
        query = f'site:linkedin.com/company "{config["name"].lower()}" "{region}"'
        try:
            keyword = config["keywords"][0].replace(" ", "+")
            search_url = f"https://www.google.com/search?q={quote(query)}&num=20"
            req = Request(search_url, headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"})
            html = urlopen(req, timeout=15).read().decode("utf-8", errors="replace")

            # Extract LinkedIn company URLs
            linkedin_matches = re.findall(r'https?://(?:www\.)?linkedin\.com/company/([^/"\s&]+)', html)
            for slug in linkedin_matches:
                slug = slug.strip().lower()
                if slug in ('in', 'jobs', 'feed', 'pub', 'mynetwork'):
                    continue
                leads.append({
                    "company_name": slug.replace("-", " ").title(),
                    "linkedin_url": f"https://www.linkedin.com/company/{slug}",
                    "source": f"Google Search ({region})",
                    "source_type": "google_linkedin",
                    "industry": industry,
                    "country": region,
                    "found_at": datetime.now(timezone.utc).isoformat(),
                })
            time.sleep(1)  # Be respectful
        except Exception as e:
            print(f"     ⚠️ Search for {region}: {e}")

    print(f"     Found {len(leads)} LinkedIn companies")
    return leads


# ── AI ENGINE: Classify, Score, Enrich ────────────────────
def ai_score_leads(leads: list[dict], industry: str) -> list[dict]:
    """Use DeepSeek to score leads by purchase intent and enrich with buyer profile."""
    if not leads:
        return []

    config = INDUSTRIES[industry]
    # Batch process: 10 leads per API call
    scored = []
    batch_size = 10

    for i in range(0, len(leads), batch_size):
        batch = leads[i:i + batch_size]
        print(f"  🤖 AI scoring batch {i // batch_size + 1} ({len(batch)} leads)...")

        companies_text = "\n".join([
            f"{j+1}. {l['company_name']} — {l.get('country', 'Unknown')} — Source: {l.get('source', '')}"
            for j, l in enumerate(batch)
        ])

        prompt = f"""Score these companies as potential buyers for Chinese {config['name']} exporters.

Buyer profiles: {'; '.join(config['buyer_profiles'])}
Target regions: {', '.join(config['target_regions'])}

Companies to evaluate:
{companies_text}

Return ONLY valid JSON array (no markdown). Each item must have:
[
  {{"idx": 1, "company_name": "Exact Name", "buyer_type": "e.g. Automotive Manufacturer", "relevance": 85, "size_hint": "Enterprise (>1000)", "purchase_urgency": "High/Medium/Low", "why": "1 sentence why they would buy from China"}},
  ...
]

relevance: 0-100. 80+ = highly likely buyer. Below 30 = probably not a buyer.
Be honest — don't inflate scores."""

        try:
            result = call_deepseek([{"role": "user", "content": prompt}], max_tokens=2000, temperature=0.3)
            result = result.replace("```json", "").replace("```", "").strip()
            scored_batch = json.loads(result)

            for item in scored_batch:
                idx = item.get("idx", 1) - 1
                if 0 <= idx < len(batch):
                    batch[idx]["buyer_type"] = item.get("buyer_type", "")
                    batch[idx]["relevance"] = item.get("relevance", 50)
                    batch[idx]["size_hint"] = item.get("size_hint", "")
                    batch[idx]["purchase_urgency"] = item.get("purchase_urgency", "Low")
                    batch[idx]["ai_why"] = item.get("why", "")
                    scored.append(batch[idx])
            time.sleep(0.5)
        except Exception as e:
            print(f"     ⚠️ AI failed: {e}")
            # Keep leads with default scores
            for l in batch:
                l["relevance"] = 30
            scored.extend(batch)

    return scored


# ── DEDUP ─────────────────────────────────────────────────
def dedup_leads(leads: list[dict], seen: set) -> tuple[list[dict], set]:
    """Remove duplicates based on company name similarity and domain tracking."""
    unique = []
    new_seen = seen.copy()

    for lead in leads:
        domain = lead.get("company_name", "").lower().strip()
        if not domain:
            continue
        # Simple dedup by normalized name
        normalized = re.sub(r'[^a-z0-9]', '', domain)
        if len(normalized) < 3:
            continue
        if normalized not in new_seen:
            new_seen.add(normalized)
            unique.append(lead)

    return unique, new_seen


# ── EXCEL GENERATION ──────────────────────────────────────
def generate_excel(leads: list[dict], industry: str):
    """Generate .xlsx file using openpyxl for all leads in this run."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️ openpyxl not installed — skipping Excel")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"{INDUSTRIES[industry]['name'][:20]}"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Color fills for relevance
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    mid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Red

    headers = [
        "Company Name", "Buyer Type", "Relevance", "Urgency",
        "Size", "Country", "Source", "LinkedIn URL", "Why Buy"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows — sorted by relevance descending
    sorted_leads = sorted(leads, key=lambda x: x.get("relevance", 0), reverse=True)

    for row_idx, lead in enumerate(sorted_leads, 2):
        values = [
            lead.get("company_name", ""),
            lead.get("buyer_type", ""),
            lead.get("relevance", 0),
            lead.get("purchase_urgency", "Low"),
            lead.get("size_hint", ""),
            lead.get("country", "Unknown"),
            lead.get("source", ""),
            lead.get("linkedin_url", ""),
            lead.get("ai_why", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color-code relevance
            if col == 3:
                if val >= 70:
                    cell.fill = high_fill
                elif val >= 40:
                    cell.fill = mid_fill
                else:
                    cell.fill = low_fill

    # Column widths
    col_widths = [35, 30, 10, 12, 15, 15, 25, 40, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Freeze header
    ws.freeze_panes = "A2"

    # Save
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"leads_{industry}_{date_str}.xlsx"
    filepath = OUTPUT_DIR / filename
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(filepath))
    print(f"  📊 Excel: {filepath} ({len(sorted_leads)} leads)")
    return filepath


# ── HTML DASHBOARD ────────────────────────────────────────
def generate_html(all_leads: list[dict]):
    """Generate a beautiful HTML dashboard for all industries."""
    # Group by industry
    groups = {}
    for lead in all_leads:
        ind = lead.get("industry", "other")
        groups.setdefault(ind, []).append(lead)

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    total = len(all_leads)
    high = sum(1 for l in all_leads if l.get("relevance", 0) >= 70)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>B2B Lead Miner — Overseas Buyer Intelligence</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#333}}
.header{{background:linear-gradient(135deg,#1a1a2e,#16213e);color:#fff;padding:30px 40px}}
.header h1{{font-size:1.8em;margin-bottom:6px}}
.header p{{color:#8892b0;font-size:.95em}}
.stats{{display:flex;gap:20px;padding:20px 40px;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
.stat{{text-align:center;flex:1}}
.stat-num{{font-size:2em;font-weight:700;color:#2B5797}}
.stat-label{{font-size:.8em;color:#888;text-transform:uppercase}}
.container{{max-width:1400px;margin:20px auto;padding:0 40px}}
.section{{margin-bottom:30px}}
.section h2{{font-size:1.2em;color:#2B5797;margin-bottom:12px;padding-bottom:8px;border-bottom:2px solid #2B5797}}
table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
th{{background:#2B5797;color:#fff;padding:12px 14px;text-align:left;font-size:.85em;text-transform:uppercase}}
td{{padding:10px 14px;border-bottom:1px solid #e8e8e8;font-size:.88em}}
tr:hover{{background:#f7f9fc}}
.score-high{{background:#d4edda;color:#155724;padding:4px 10px;border-radius:4px;font-weight:600}}
.score-mid{{background:#fff3cd;color:#856404;padding:4px 10px;border-radius:4px;font-weight:600}}
.score-low{{background:#f8d7da;color:#721c24;padding:4px 10px;border-radius:4px;font-weight:600}}
.urgency-high{{color:#dc3545;font-weight:600}}
.urgency-medium{{color:#fd7e14;font-weight:600}}
.urgency-low{{color:#6c757d}}
footer{{text-align:center;color:#999;padding:30px;font-size:.8em}}
</style>
</head>
<body>
<div class="header">
<h1>🌍 B2B Lead Miner</h1>
<p>Overseas Buyer Intelligence for Chinese Robotics & Energy Storage Exporters</p>
</div>
<div class="stats">
<div class="stat"><div class="stat-num">{total}</div><div class="stat-label">Total Leads</div></div>
<div class="stat"><div class="stat-num">{high}</div><div class="stat-label">High Relevance (≥70)</div></div>
<div class="stat"><div class="stat-num">{len(groups)}</div><div class="stat-label">Industries</div></div>
<div class="stat"><div class="stat-num">{now}</div><div class="stat-label">Last Updated</div></div>
</div>
<div class="container">
"""

    for industry, leads in groups.items():
        ind_name = INDUSTRIES.get(industry, {}).get("name", industry)
        sorted_leads = sorted(leads, key=lambda x: x.get("relevance", 0), reverse=True)
        html += f'<div class="section"><h2>🤖 {ind_name} — {len(sorted_leads)} leads</h2>'
        html += '<table><tr><th>Company</th><th>Type</th><th>Score</th><th>Urgency</th><th>Size</th><th>Country</th><th>Source</th><th>LinkedIn</th></tr>'

        for l in sorted_leads[:50]:  # Top 50 per industry
            score = l.get("relevance", 0)
            score_class = "score-high" if score >= 70 else ("score-mid" if score >= 40 else "score-low")
            urgency = l.get("purchase_urgency", "Low").lower()
            urgency_class = f"urgency-{urgency}" if urgency in ("high","medium","low") else "urgency-low"
            linkedin = l.get("linkedin_url", "")
            linkedin_link = f'<a href="{linkedin}" target="_blank">🔗</a>' if linkedin else "—"

            html += f"""<tr>
<td><strong>{l.get("company_name","")}</strong></td>
<td>{l.get("buyer_type","")}</td>
<td><span class="{score_class}">{score}</span></td>
<td><span class="{urgency_class}">{l.get("purchase_urgency","Low")}</span></td>
<td>{l.get("size_hint","")}</td>
<td>{l.get("country","")}</td>
<td>{l.get("source","")}</td>
<td>{linkedin_link}</td>
</tr>"""

        html += "</table></div>"

    html += f"""
</div>
<footer>B2B Lead Miner — Auto-generated every 6 hours. Powered by AI + GitHub Actions.</footer>
</body>
</html>"""

    filepath = OUTPUT_DIR / "dashboard.html"
    filepath.write_text(html)
    print(f"  🌐 HTML Dashboard: {filepath}")


# ── MAIN ─────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  🌍 B2B Lead Miner")
    print(f"  {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * 60)

    state = load_state()
    seen = load_seen()
    all_leads = []

    # For each industry, run all data sources
    for industry_key in ["robotics", "energy_storage"]:
        print(f"\n┌─ {INDUSTRIES[industry_key]['name']} ─┐")
        industry_leads = []

        # 1. Trade show scraping
        print("│ Trade Shows:")
        for show_id, show in TRADE_SHOWS.items():
            if industry_key in show["industries"]:
                leads = scrape_trade_show(show, industry_key)
                industry_leads.extend(leads)
                time.sleep(1)

        # 2. Google → LinkedIn
        print("│ Google Search → LinkedIn:")
        linkedin_leads = scrape_google_linkedin(industry_key)
        industry_leads.extend(linkedin_leads)

        # 3. AI scoring
        print(f"│ AI Scoring ({len(industry_leads)} raw leads)...")
        if industry_leads:
            industry_leads = ai_score_leads(industry_leads, industry_key)

        # 4. Dedup
        industry_leads, seen = dedup_leads(industry_leads, seen)
        print(f"│ After dedup: {len(industry_leads)} unique new leads")

        # 5. Generate Excel
        if industry_leads:
            generate_excel(industry_leads, industry_key)

        all_leads.extend(industry_leads)

    # Generate combined HTML
    if all_leads:
        generate_html(all_leads)

    # Save state
    state["total_leads"] += len(all_leads)
    save_state(state)
    save_seen(seen)

    print(f"\n✨ Done. {len(all_leads)} new leads this run. Total leads mined: {state['total_leads']}")


if __name__ == "__main__":
    main()

